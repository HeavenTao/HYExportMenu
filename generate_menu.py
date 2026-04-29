#!/usr/bin/env python3
"""
根据menu.json生成Excel文件，参照模板.xlsx的格式
支持单元格合并，过滤无url的行，文本左对齐垂直居中
"""
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy


# 全局对齐样式：水平居左，垂直居中
CELL_ALIGNMENT = Alignment(horizontal='left', vertical='center')


def load_menu_data(filepath):
    """加载menu.json数据"""
    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)


def copy_cell_style(src_cell, dst_cell):
    """复制单元格样式，并覆盖对齐方式"""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        # 使用自定义对齐方式
        dst_cell.alignment = CELL_ALIGNMENT
        dst_cell.border = copy(src_cell.border)
        dst_cell.number_format = copy(src_cell.number_format)


def copy_worksheet_format(src_ws, dst_ws):
    """复制工作表的格式（列宽、行高等）"""
    # 复制列宽
    for col_letter, col_dim in src_ws.column_dimensions.items():
        if col_dim.width:
            dst_ws.column_dimensions[col_letter].width = col_dim.width
    
    # 复制行高
    for row_num, row_dim in src_ws.row_dimensions.items():
        if row_dim.height:
            dst_ws.row_dimensions[row_num].height = row_dim.height


def generate_menu_data(menu_data):
    """
    根据menu.json生成菜单数据，过滤掉没有url地址的行
    返回:
        rows: 列表，每个元素是 (一级菜单名, 二级菜单名, 三级菜单名, 地址)
        first_merges: 一级菜单合并范围 [(菜单名, 起始行, 结束行), ...]
        second_merges: 二级菜单合并范围 [(菜单名, 起始行, 结束行), ...]
    """
    rows = []
    
    for first_level in menu_data:
        first_name = first_level.get('Name', '')
        second_menus = first_level.get('SubMenus', [])
        
        for second_level in second_menus:
            second_name = second_level.get('Name', '')
            third_menus = second_level.get('SubMenus', [])
            
            for third_level in third_menus:
                third_name = third_level.get('Name', '')
                url = third_level.get('NavigateUrl', '')
                
                # 只保留有url的行（非空字符串且非None）
                if url and url.strip():
                    rows.append((first_name, second_name, third_name, url))
    
    # 计算合并范围
    first_merges = []
    second_merges = []
    
    if not rows:
        return rows, first_merges, second_merges
    
    # 计算一级菜单合并范围
    current_first = rows[0][0]
    first_start = 3  # Excel行号从3开始（1=说明，2=表头）
    
    for idx, row in enumerate(rows):
        if row[0] != current_first:
            first_end = idx + 2  # 上一行的行号
            first_merges.append((current_first, first_start, first_end))
            current_first = row[0]
            first_start = idx + 3  # 当前行的行号
    
    # 最后一个一级菜单
    first_merges.append((current_first, first_start, len(rows) + 2))
    
    # 计算二级菜单合并范围
    current_second = rows[0][1]
    current_first_for_second = rows[0][0]
    second_start = 3
    
    for idx, row in enumerate(rows):
        # 当二级菜单变化，或一级菜单变化（同名二级菜单在不同一级下）
        if row[1] != current_second or row[0] != current_first_for_second:
            second_end = idx + 2
            second_merges.append((current_second, second_start, second_end))
            current_second = row[1]
            current_first_for_second = row[0]
            second_start = idx + 3
    
    # 最后一个二级菜单
    second_merges.append((current_second, second_start, len(rows) + 2))
    
    return rows, first_merges, second_merges


def set_cell_value(ws, row, col, value):
    """设置单元格值和对齐方式"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = CELL_ALIGNMENT
    return cell


def create_excel(menu_data, template_path, output_path):
    """创建Excel文件"""
    # 加载模板
    template_wb = openpyxl.load_workbook(template_path)
    template_ws = template_wb.active
    
    # 创建新的工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = template_ws.title
    
    # 复制工作表格式
    copy_worksheet_format(template_ws, ws)
    
    # 复制说明行和表头行（处理合并单元格）
    for row_idx in range(1, 3):
        for col_idx in range(1, template_ws.max_column + 1):
            src_cell = template_ws.cell(row=row_idx, column=col_idx)
            dst_cell = ws.cell(row=row_idx, column=col_idx)
            if not isinstance(dst_cell, MergedCell):
                dst_cell.value = src_cell.value
                copy_cell_style(src_cell, dst_cell)
    
    # 复制说明行的合并单元格 A1:C1
    ws.merge_cells('A1:C1')
    
    # 生成菜单数据
    rows, first_merges, second_merges = generate_menu_data(menu_data)
    
    # 写入数据
    for idx, (first, second, third, url) in enumerate(rows, start=3):
        set_cell_value(ws, idx, 1, first)
        set_cell_value(ws, idx, 2, second)
        set_cell_value(ws, idx, 3, third)
        set_cell_value(ws, idx, 4, url)
    
    # 一级菜单合并单元格
    for first_name, start_row, end_row in first_merges:
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
    
    # 二级菜单合并单元格
    for second_name, start_row, end_row in second_merges:
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
    
    # 如果有Sheet2，复制它
    if 'Sheet2' in template_wb.sheetnames:
        template_ws2 = template_wb['Sheet2']
        ws2 = wb.create_sheet(title='Sheet2')
        copy_worksheet_format(template_ws2, ws2)
        for row_idx in range(1, template_ws2.max_row + 1):
            for col_idx in range(1, template_ws2.max_column + 1):
                src_cell = template_ws2.cell(row=row_idx, column=col_idx)
                dst_cell = ws2.cell(row=row_idx, column=col_idx)
                if not isinstance(dst_cell, MergedCell):
                    dst_cell.value = src_cell.value
                    copy_cell_style(src_cell, dst_cell)
    
    # 保存文件
    wb.save(output_path)
    print(f"Excel文件已生成: {output_path}")
    print(f"总共 {len(rows)} 行菜单数据")
    print(f"一级菜单合并: {len(first_merges)} 处")
    print(f"二级菜单合并: {len(second_merges)} 处")


def main():
    menu_data = load_menu_data('menu.json')
    create_excel(menu_data, '模板.xlsx', '菜单导出.xlsx')


if __name__ == '__main__':
    main()
